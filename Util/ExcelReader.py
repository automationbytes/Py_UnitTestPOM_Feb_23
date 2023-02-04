'''
openpyxl - lightweighted
            close comparison to apache poi in java
pandas - data frames

pip install openpyxl
'''

import openpyxl
def readExcel(location,label,header):
    workbook = openpyxl.load_workbook(location,read_only=True)
    print(workbook.sheetnames)
    #sheet = workbook.get_sheet_by_name("DevopsUniv")
    sheet =workbook["Sauce"]

    noofRows = sheet.max_row
    noofCols = sheet.max_column
    print(noofRows)
    print(noofCols)

    Label = []
    Header = []

    for i in range(1,noofRows+1):
        Label.append(sheet.cell(i,1).value)
        if(Label[i-1] == label):
            for j in range(1, noofCols + 1):
                Header.append(sheet.cell(1, j).value)
                if(Header[j-1] == header):
                    val = sheet.cell(i,j).value
                    break
            break
    return val
#print(Label)
#print(Header)
